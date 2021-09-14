from behave import *
import time
import random
from datetime import datetime, timedelta, timezone
from dateutil import tz
from behave.api.async_step import async_run_until_complete
from openpyxl import load_workbook, Workbook
from base.main_functions import GetRequest, get_token, find_button, prev_current_date
from behave.api.async_step import async_run_until_complete
from base.main_functions import random_filter_generator_with_none
from base.sql_functions import (
    pgsql_select,
    decode_request,
    encode_request,
    pgsql_insert,
    pgsql_del
)
from base.db_interactions.accounting_system import UserBill, HistoryUserBill, \
    Transaction, UserBillType, HistoryCompanyBill, CompanyBill, User, Entry, UserMainData, Process, Broker, AccountType
from base.db_interactions.reconciliation import UserPropaccounts, ReconciliationUserPropaccounts



@step("delete manager_user bills: account and withdrawal and UserMainData")
def step_impl(context):
    HistoryUserBill.filter(bill_id__in__name=['Account', 'Withdrawal']).delete()
    userbills = UserBill.filter(user_id__hr_id=context.custom_config['manager_id']['hr_id'], bill_id__in__name=['Account', 'Withdrawal'])
    Transaction.filter(user_bill_id__in=userbills).delete()
    userbills.delete()
    UserMainData.filter(user_id__hr_id=context.custom_config['manager_id']['hr_id']).delete()

@step("create manager_user bills: account and withdrawal")
def step_impl(context):
    user_hr_id = context.custom_config['manager_id']['hr_id']
    bill_types = UserBillType.filter(name__in=['Account', 'Withdrawal'])

    context.user_bills = {}
    for bill_type in bill_types:
        amount = 0
        user_bill = UserBill.create(
            user_id__hr_id=user_hr_id,
            amount=amount,
            bill_id=bill_type
        )
        context.user_bills[bill_type.name] = user_bill.id
        now = datetime.now()
        HistoryUserBill.create(
            model_id=user_bill.id,
            history_date=str(now.replace(day=1, month=now.month-3)),
            history_type='+',
            history_created=str(now),
            amount=amount,
            bill_id=bill_type,
            user_id__hr_id=user_hr_id
        )
    history_comp_bill = HistoryCompanyBill.filter(
        model_id__name='Company Daily Net'
    ).sorted_by('history_date')[0]

    history_comp_bill.history_date = (datetime.now() - timedelta(days=90))
    history_comp_bill.save()

    context.company_bill = CompanyBill.get(name='Company Daily Net').id

def give_date(day, month):
    given_date = datetime.now().replace(hour=12)
    given_month = given_date.month
    if month == 'before last':
        given_month -= 2
    elif month == 'last':
        given_month -= 1
    return str(given_date.replace(day=int(day), month=given_month))

@step("create default ENTRIES with parameters")
def step_impl(context):
    session = context.super_user
    url = context.custom_config["host"] + 'api/accounting_system/entry/'

    for row in context.table:
        request_body = {
            'transaction_out.user_bill': '',
            'transaction_out.company_bill': '',
            'transaction_in.user_bill': '',
            'transaction_in.company_bill': '',
            'entry.date_to_execute': give_date(row['day'], row['month']),
            'entry.description': 'autotest',
            'transaction_common.amount_usd': '',
            'transaction_common.description': 'autotest',
            'csrfmiddlewaretoken': get_token(session, url),
        }

        if '-' in row['amount']:
            request_body['transaction_out.user_bill'] = context.user_bills[row['user_bill']]
            request_body['transaction_in.company_bill'] = context.company_bill
            request_body['transaction_common.amount_usd'] = float(row['amount'][1:])
        else:
            request_body['transaction_in.user_bill'] = context.user_bills[row['user_bill']]
            request_body['transaction_out.company_bill'] = context.company_bill
            request_body['transaction_common.amount_usd'] = float(row['amount'])

        response = session.post(url, data=request_body, headers={"Referer": url})

        time.sleep(1)

        context.table_again = context.table

@step("create UserMainData for manager user")
def step_impl(context):
    for row in context.table:
        datum = give_date(28, row['month'])
        UserMainData.create(
            services_total=int(row['services']),
            compensations_total=int(row['compensations']),
            office_fees=int(row['fees']),
            prev_month_net=int(row['prev_month']),
            total_net_month=int(row['total_net']),
            deadline=2500,
            payout_rate=0.5,
            change_plus_minus=int(row['plus_minus']),
            zp_cash=int(row['zp_cash']),
            company_cash=int(row['company_cash']),
            social=int(row['social']),
            withdrawal=int(row['withdrawal']),
            effective_date=datum,
            user_id__hr_id=context.custom_config['manager_id']['hr_id']
        )

@step("create expected result for Account")
def step_impl(context):
    now = datetime.now()
    prev_month = now.replace(day=1) - timedelta(days=1)
    prevprev_month = prev_month.replace(day=1) - timedelta(days=1)
    prevprevprev_month = prevprev_month.replace(day=1) - timedelta(days=1)

    month_index = {
        'before last': 1,
        'last': 2,
        'this': 3
    }

    exp_account_result = [
        {
            'date': prevprevprev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'amount': 0,
            'changes': 0
        },
        {
            'date': prevprev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'amount': 0,
            'changes': 0
        },
        {
            'date': prev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'amount': 0,
            'changes': 0
        },
        {
            'date': now.strftime('%Y-%m-%d'),
            'amount': 0,
            'changes': 0
        }
    ]

    for row in context.table_again:
        if row['user_bill'] == 'Account':
            exp_account_result[month_index[row['month']]]['changes'] += float(row['amount'])

    for i in range(1, len(exp_account_result)):
        exp_account_result[i]['amount'] = round(exp_account_result[i]['changes'] + exp_account_result[i - 1]['amount'], 4)
    exp_account_result.reverse()

    context.exp_account_result = exp_account_result

@step("create expected result for Withdrawal")
def step_impl(context):
    now = datetime.now()
    prev_month = now.replace(day=1) - timedelta(days=1)
    prevprev_month = prev_month.replace(day=1) - timedelta(days=1)
    prevprevprev_month = prevprev_month.replace(day=1) - timedelta(days=1)

    month_index = {
        'before last': 1,
        'last': 2,
        'this': 3
    }

    exp_withdrawal_result = [
        {
            'date': prevprevprev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'left': 0,
            'amount': 0,
            'all_time_total': 0
        },
        {
            'date': prevprev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'left': 0,
            'amount': 0,
            'all_time_total': 0
        },
        {
            'date': prev_month.strftime('%Y-%m-%dT23:59:59.999999'),
            'left': 0,
            'amount': 0,
            'all_time_total': 0
        },
        {
            'date': now.strftime('%Y-%m-%d'),
            'left': 0,
            'amount': 0,
            'all_time_total': 0
        }
    ]

    for row in context.table_again:
        if row['user_bill'] == 'Withdrawal':
            exp_withdrawal_result[month_index[row['month']]]['left'] += float(row['amount'])
            if '-' in row['amount']:
                exp_withdrawal_result[month_index[row['month']]]['amount'] += float(row['amount'])

    for i in range(1, len(exp_withdrawal_result)):
        exp_withdrawal_result[i]['left'] += exp_withdrawal_result[i-1]['left']
        exp_withdrawal_result[i]['all_time_total'] = round(exp_withdrawal_result[i]['amount'] + exp_withdrawal_result[i-1]['all_time_total'], 4)

    exp_withdrawal_result.reverse()
    context.exp_withdrawal_result = exp_withdrawal_result

@step("compare manager actual and expected {bill_type} results")
def step_impl(context, bill_type):
    bill_type_id = UserBillType.get(name=bill_type).id

    session = context.manager_user
    result = session.get(context.custom_config["host"] + f"api/accounting_system/accounting/{bill_type_id}/")
    actual_result = result.json()
    actual_result[0]['date'] = actual_result[0]['date'].split('T')[0]

    with open('./xxx.txt', 'a') as file:
        file.write(str(actual_result) + '\n')
    with open('./xxx.txt', 'a') as file:
        file.write(str(context.exp_account_result) + '\n')
    if bill_type == 'Account':
        assert actual_result == context.exp_account_result
    elif bill_type == 'Withdrawal':
        assert actual_result == context.exp_withdrawal_result

@step("compare by risk actual and expected {bill_type} results")
def step_impl(context, bill_type):
    bill_type_id = UserBillType.get(name=bill_type).id
    user_id = User.get(hr_id = context.custom_config['manager_id']['hr_id']).id

    session = context.super_user
    result = session.get(context.custom_config["host"] + f"api/accounting_system/accounting/{bill_type_id}/?user={user_id}")
    actual_result = result.json()
    actual_result[0]['date'] = actual_result[0]['date'].split('T')[0]

    if bill_type == 'Account':
        assert actual_result == context.exp_account_result
    elif bill_type == 'Withdrawal':
        assert actual_result == context.exp_withdrawal_result

def create_transaction_list(user, bill_type, date_from, date_to, broker=[]):
    user_bill = UserBill.get(
        bill_id__name=bill_type,
        user_id=user.id
    )
    if not user_bill:
        return []

    enr = Entry.filter(
        date_to_execute__gte=str(date_from),
        date_to_execute__lt=str(date_to)
    )

    transactions = Transaction.filter(
        user_bill_id=user_bill.id,
        entry_id__in=enr
    )

    list_transactions = []

    for transaction in transactions:

        if transaction.account_type_id:
            acc_type_id = AccountType.get(id=transaction.account_type_id)
        else:
            acc_type_id = None

        if broker != [] and (acc_type_id == None or acc_type_id.broker_id not in broker):
            continue

        if transaction.initiated_user_id:
            initiated_user_obj = User.get(id=transaction.initiated_user_id)
            initiated_user = {
                    "id": initiated_user_obj.id,
                    "first_name": initiated_user_obj.first_name,
                    "last_name": initiated_user_obj.last_name,
                    "patronymic": initiated_user_obj.patronymic,
                    "hr_id": initiated_user_obj.hr_id
                }

        else:
            initiated_user = None

        if transaction.initiated_process_id:
            init_proc = {
                'name': Process.get(id=transaction.initiated_process_id).name
            }
        else:
            init_proc = None

        entry = Entry.get(id=transaction.entry_id)

        list_transactions.append(
            {
                "id": transaction.id,
                "side": transaction.side,
                "user_bill": {
                    "id": user_bill.id,
                    "user": {
                        "id": user.id,
                        "first_name": user.first_name,
                        "last_name": user.last_name,
                        "patronymic": user.patronymic,
                        "hr_id": user.hr_id
                    },
                    "bill": {
                        "id": user_bill.bill_id,
                        "name": bill_type
                    }
                },
                "company_bill": None,
                "amount": '{:.4f}'.format(transaction.amount),
                "currency": transaction.currency,
                "rate_to_usd": '{:.8f}'.format(transaction.rate_to_usd),
                "amount_usd": '{:.4f}'.format(transaction.amount_usd),
                "created": transaction.created.astimezone(tz.gettz('Europe/Kiev')).strftime("%Y-%m-%dT%H:%M:%S.%f"),
                "initiated_user": initiated_user,
                "initiated_process": init_proc,
                "status": "Applied",
                "description": transaction.description,
                "entry": {
                    "id": entry.id,
                    "date_to_execute": entry.date_to_execute.astimezone(tz.gettz('Europe/Kiev')).strftime("%Y-%m-%dT%H:%M:%S.%f")
                }
            }
        )
    return list_transactions

@step("get transactions data from db - bill:{bill_type}, month:{month}")
def step_impl(context, bill_type, month):
    datum_1 = datetime.fromisoformat(give_date(1, month))
    datum_2 = datum_1.replace(month=datum_1.month+1)
    user = User.get(hr_id=context.custom_config['manager_id']['hr_id'])

    context.expected_transactions = create_transaction_list(user, bill_type, datum_1.date(), datum_2.date())

@step("compare by risk and manager actual and expected {bill_type} results for {month} month")
def step_impl(context, bill_type, month):
    datum_1 = datetime.fromisoformat(give_date(1, month))
    bill_type_id = UserBillType.get(name=bill_type).id
    user_id = User.get(hr_id=context.custom_config['manager_id']['hr_id']).id

    session = context.super_user
    result = session.get(
        context.custom_config["host"] +
        f"api/accounting_system/accounting/{bill_type_id}/{datum_1.year}/{datum_1.month}/?user={user_id}"
    )
    actual_risk_result = result.json()
    session = context.manager_user
    result = session.get(
        context.custom_config["host"] +
        f"api/accounting_system/accounting/{bill_type_id}/{datum_1.year}/{datum_1.month}/"
    )
    actual_manager_result = result.json()
    assert actual_risk_result == actual_manager_result


    for trans in context.expected_transactions:
        with open('./xxx.txt', 'a') as file:
            file.write(str(actual_risk_result) + '\n')
        with open('./xxx.txt', 'a') as file:
            file.write(str(trans) + '\n')
        assert trans in actual_risk_result

@step("create random MANAGER report url and get actual result")
def step_impl(context):
    datum = datetime.now().replace(day=1) - timedelta(days=random.randint(0, 1))
    context.api_datum = datum.replace(
        day=random.randint(1, 28)
    )

    field_list = [
        'services_total',
        'compensations_total',
        'office_fees',
        'prev_month_net',
        'total_net_month',
        'change_plus_minus',
        'zp_cash',
        'company_cash',
        'social',
        'withdrawal',
    ]
    url_field, context.new_field_list = random_filter_generator_with_none(field_list, 'field')

    bill_list = [
        'Account',
        'Withdrawal',
        # 'Cash hub',
        # 'Current Net balance'
    ]
    bill_list_id = [bill.id for bill in UserBillType.filter(name__in=bill_list)]
    url_bill, context.new_bill_list = random_filter_generator_with_none(bill_list_id, 'account')

    url = context.custom_config["host"] + \
          f"api/accounting_system/report/?{url_field+url_bill}date={context.api_datum.date()}"

    session = context.manager_user
    response = session.get(url)
    context.actual_result = response.json()

@step("get expected MANAGER report result")
def step_impl(context):
    user_id = context.custom_config['manager_id']['user_id']

    usermaindata = UserMainData.filter(
        user_id=user_id,
        effective_date__lt=str(context.api_datum)
    ).sorted_by('effective_date', desc=True)
    fields = {field: float(getattr(usermaindata[0], field)) for field in context.new_field_list}

    context.api_datum = context.api_datum.replace(
        hour=23,
        minute=59
    ).astimezone(timezone.utc)
    accounts = {}
    for bill in context.new_bill_list:
        userbill = HistoryUserBill.filter(
            user_id=user_id,
            bill_id=bill,
            history_date__lt=str(context.api_datum)
        ).sorted_by('history_date', desc=True)
        accounts[UserBillType.get(id=bill).name] = float(userbill[0].amount)

    if accounts == fields == {}:
        context.expected_result = {
            "errors": {

                "field_account": "Field or Account is required"
            },
            "status_code": 400
        }
    else:
        context.expected_result = {
            user_id: {
                'accounts': accounts,
                'fields': fields
            }
        }

@step("compare actual and expected result of MANAGER report")
def step_impl(context):
    assert context.expected_result == context.actual_result


@step("get actual MANAGER journal entries with random data")
def step_impl(context):
    context.date_from = datetime.now().replace(hour=0, minute=0, microsecond=0) - timedelta(days=random.randint(20, 30))
    context.date_to = datetime.now().replace(hour=23, minute=59, microsecond=59) - timedelta(days=random.randint(0, 19))
    bill_list = [
        'Account',
        'Withdrawal',
        # 'Cash hub',
        # 'Current Net balance'
    ]

    bill_list_id = [bill.id for bill in UserBillType.filter(name__in=bill_list)]
    url_bill, context.new_bill_list = random_filter_generator_with_none(bill_list_id, 'account')
    context.side = random.randint(0, 1)
    url_broker, context.new_broker_list = random_filter_generator_with_none([1, 3], 'acc_broker')

    url = context.custom_config["host"] + \
          f"api/accounting_system/transactions/?{url_broker + url_bill}transaction_side={context.side}" \
          f"&date_from={context.date_from.date()}T00:00:00&date_to={context.date_to.date()}T23:59:59.99"
    session = context.manager_user
    response = session.get(url)

    context.actual_result = response.json()['data']

@step("get expected MANAGER journal entries with random data")
def step_impl(context):
    context.expected_result = []
    user = User.get(hr_id=context.custom_config['manager_id']['hr_id'])

    for bill in context.new_bill_list:

        transactions = create_transaction_list(
            user,
            UserBillType.get(id=bill).name,
            context.date_from.astimezone(timezone.utc),
            context.date_to.astimezone(timezone.utc),
            context.new_broker_list
        )
        transactions = [transaction for transaction in transactions if transaction['side'] == context.side]
        context.expected_result.extend(transactions)

@step("compare actual and expected result of MANAGER journal entries")
def step_impl(context):
    for transaction in context.actual_result:
        assert transaction in context.expected_result



@step("create UserPropaccounts models of user {user} for {day} day of prev month")
def step_impl(context, user, day):
    try:
        UserPropaccounts.all().delete()
    except:
        pass

    month = prev_current_date()['prev_month']
    user_accounts = ReconciliationUserPropaccounts.filter(user_id=int(user))
    expected_result = {}

    for index in range(len(context.table.rows)):
        row = context.table[index]
        current_day = int(day) + index
        expected_result[current_day] = {
            'daily_gross': 0,
            'daily_adj_net': 0,
            'daily_unreal': 0,
            'month_gross': 0,
            'month_unreal': 0,
            'month_adj_net': 0,
        }
        for account in user_accounts:


            local_dict = expected_result[current_day].copy()
            for key, value in row.items():
                local_dict[key] = float(value) + index
                expected_result[current_day][key] += float(value) + index
                expected_result[current_day][key] = round(expected_result[current_day][key], 4)

            UserPropaccounts.create(
                account=account.account,
                effective_date=datetime.now().replace(
                    day=current_day,
                    month=month
                ),
                user_id=user,
                account_type_id=account.account_type_id,
                **local_dict
            )

    context.expected_result = expected_result

@step("get actual result from api of user {user} for {day} day of prev month")
def step_impl(context, user, day):
    month = prev_current_date()['prev_month']
    session = context.fin_user
    url = context.custom_config["host"] + f'api/accounting_system/report/?user[]={user}' \
                                          f'&field[]=month_adj_net&field[]=month_gross' \
                                          f'&field[]=month_unreal&field[]=daily_unreal' \
                                          f'&field[]=daily_gross&field[]=daily_adj_net' \
                                          '&date={}'

    actual_result = {}
    for i in range(len(context.expected_result)):
        datum = datetime.now().replace(
            day=int(day)+i,
            month=month
        )
        response = session.get(url.format(datum.date()))
        # with open('./xxx.txt', 'a') as file:
        #     file.write(str(url.format(datum.date())) + '\n')
        actual_result[int(day)+i] = response.json()[user]['fields']

    context.actual_result = actual_result

@step("[Report with fields] compare expected and actual result")
def step_impl(context):
    with open('./xxx.txt', 'a') as file:
        file.write(str(context.actual_result ) + '\n')
    with open('./xxx.txt', 'a') as file:
        file.write(str(context.expected_result) + '\n')
    assert context.actual_result == context.expected_result


@step('upload manager\'s account data through Riskbot')
@async_run_until_complete
async def send_fees_to_riskbot(context):
    archives_list = [
        'base/data_set/prev_acc_fees.7z',
        'base/data_set/next_acc_fees.7z'
    ]
    for path in archives_list:
        async with context.tele_user.conversation(context.custom_config["risk_bot"]) as conv:
            await conv.send_message('/start')
            message = await conv.get_response()
            button = find_button([message], 'Загрузить данные по account')
            await button.click()
            await conv.get_response()
            await conv.send_file(path)
            answer = await conv.get_response()

            assert answer.text == 'Файл успешно загружен, данные для сверки отправлены'

@step("get data from accounting:{type_}")
def step_impl(context, type_):
    if type_ == "account_data":
        url = context.custom_config["host"] + "api/accounting/account_data/"
    elif type_ == "account_queue":
        url = context.custom_config["host"] + "api/accounting/payment_queue_data/"

    session = context.manager_user
    response = session.get(url)
    context.act_data = response.json()

    assert response.ok


@step("in accounting compare s expected with actual {type_}")
def step_impl(context, type_):
    if type_ == "account_data":
        expected_data = context.account_data
    elif type_ == "account_queue":
        expected_data = context.queque_accounting
    assert context.act_data == expected_data

@step("in accounting compare expected with actual {type_} file")
def step_impl(context, type_):

    wb_act = load_workbook(f'base/data_set/{type_}.xlsx')
    ws_act = wb_act.active
    values_act = list(ws_act.values)

    wb_exp = load_workbook(f'base/data_set/{type_}_template.xlsx')
    ws_exp = wb_exp.active
    values_exp = list(ws_exp.values)

    assert values_act == values_exp

@step("download {file} xlsx")
def step_impl(context, file):

    if file == 'account_data':
        url = context.custom_config["host"] + "api/accounting/data_history/xlsx_account/"
    elif file == "account_queue":
        url = context.custom_config["host"] + "api/accounting/data_history/xlsx_payment_queue/"

    session = context.manager_user
    response = session.get(url)
    assert response.ok

    with open(f'base/data_set/{file}.xlsx', 'wb') as file:
        file.write(response.content)


@step("get transactions of all users by FIN")
def step_impl(context):
    from_date = datetime.now() - timedelta(days=100)
    to_date = datetime.now()

    session = context.fin_user
    context.url = context.custom_config["host"] + "api/accounting_system/transactions/?user[]=all&account[]=all&" \
                                     f"date_from={from_date.date()}T00:00:00&date_to={to_date.date()}T23:59:59.99&page_size=1000"

    response = session.get(context.url).json()
    context.expected_data = []
    for transaction in response['data']:
        user_hr_id = transaction['user_bill']['user']['hr_id']
        first_name = transaction['user_bill']['user']['first_name']
        last_name = transaction['user_bill']['user']['last_name']
        entry_id = transaction['entry']['id']
        transaction_id = transaction['id']
        initiator = None

        if transaction['initiated_user']:
            initiator = f"{transaction['initiated_user']['hr_id']} {transaction['initiated_user']['first_name']} {transaction['initiated_user']['last_name']}"
        elif transaction['initiated_process']:
            initiator = transaction['initiated_process']['name']
        try:
            date_to_execute = datetime.strptime(transaction['entry']['date_to_execute'],
                                                '%Y-%m-%dT%H:%M:%S.%f').replace(microsecond=0)
        except:
            date_to_execute = datetime.strptime(transaction['entry']['date_to_execute'],
                                                '%Y-%m-%dT%H:%M:%S')
        date_created = datetime.strptime(transaction['created'], '%Y-%m-%dT%H:%M:%S.%f').replace(microsecond=0)
        bill_name = transaction['user_bill']['bill']['name']
        amount = float(transaction['amount']) * (2 * transaction['side'] - 1)
        status = transaction['status']
        description = transaction['description']

        context.expected_data.append([
            f'{user_hr_id} {first_name} {last_name}',
            entry_id,
            transaction_id,
            initiator,
            date_to_execute,
            date_created,
            bill_name,
            amount,
            status,
            description
        ])

@step("download xlsx file with transactions")
def step_impl(context):
    session = context.fin_user
    response = session.get(context.url+'&file_format=xlsx')
    with open(f'base/data_set/transactions.xlsx', 'wb') as file:
        file.write(response.content)
    wb = load_workbook('base/data_set/transactions.xlsx')
    ws = wb.active
    actual_data = list(ws.values)[1:]
    context.actual_data = []
    for row in actual_data:
        row_list = list(row)
        row_list[4] = row_list[4].replace(microsecond=0)
        row_list[5] = row_list[5].replace(microsecond=0)
        context.actual_data.append(row_list)

@step("compare xlsx transactions: actual with expected")
def step_impl(context):
    if context.expected_data != context.actual_data:
        context.logger.error('---------------MSW-727---------------')
        context.logger.error(str(context.expected_data))
        context.logger.error(str(context.actual_data))
    assert context.expected_data == context.actual_data
