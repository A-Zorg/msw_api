from openpyxl import Workbook
from openpyxl import load_workbook
import py7zr
import pandas
from datetime import datetime, timedelta, date
import random

def update_fees(from_data, to_data):
    for i in range(len(to_data)):
        for j in range(len(to_data[i])):
            to_data[i][j].value=from_data[i][j].value

def create_riskbot_fees():
    """------------------------create fees file for riskbot----------------------------"""
    fees = pandas.read_csv('base/data_set/fees.csv')
    rows = fees.index.stop
    wb_from = load_workbook('base/data_set/fees.xlsx')
    ws_from = wb_from.active
    wb_to = load_workbook('generator/december.xlsx')
    ws_to = wb_to.active

    from_data = ws_from['D2':f'H{rows + 1}']
    to_data = ws_to['K5':f'O{rows + 4}']
    update_fees(from_data, to_data)

    from_user_id = ws_from['B2':f'B{rows + 1}']
    to_user_id = ws_to['B5':f'B{rows + 4}']
    update_fees(from_user_id, to_user_id)

    tod_ay = datetime.today()
    month = (tod_ay-timedelta(tod_ay.day+1)).month
    year = (tod_ay - timedelta(tod_ay.day+1)).year

    ws_to['C2'] = month-1
    ws_to['C3'] = date(year, month, 11)

    wb_to.save('base/data_set/fees_riskbot.xlsx')

    """------------------------create fees archive for riskbot----------------------------"""

    with py7zr.SevenZipFile('base/data_set/fees.7z', 'w', password='123456') as archive:
        archive.writeall('base/data_set/fees_riskbot.xlsx', 'month.xlsx',)

def random_fees_creator():
    notes = [
        ['base/data_set/next_acc_fees.xlsx', ''],
        ['base/data_set/prev_acc_fees.xlsx', ''],
    ]
    for i in range(1, len(notes)+1):
        month_dict = {
            'datum' : (date.today() - timedelta(weeks=5)*i).replace(day=25),
            'd' : random.randint(500,3000),
            'e' : random.randint(-500,1500),
            'g' : random.randint(6500,13000),
            'i' : random.randint(-3500,3000),
            'j' : random.randint(-3500,3000),
            'm' : random.randint(-200,-50),
            'n' : random.randint(-200,-50),
        }
        notes[i-1][1] = month_dict
    return notes

def create_7z(data_set, hr_id):
    for ds in data_set:
        wb_to = load_workbook(ds[0])
        ws_to = wb_to.active

        ws_to['C2'] = ds[1]['datum'].month - 1
        ws_to['C3'] = ds[1]['datum']
        ws_to['B5'] = int(hr_id)
        ws_to['D5'] = ds[1]['d']
        ws_to['E5'] = ds[1]['e']
        ws_to['G5'] = ds[1]['g']
        ws_to['I5'] = ds[1]['i']
        ws_to['J5'] = ds[1]['j']
        ws_to['M5'] = ds[1]['m']
        ws_to['N5'] = ds[1]['n']

        wb_to.save(ds[0])

        zip_name = ds[0][:-4]+'7z'
        with py7zr.SevenZipFile(zip_name, 'w', password='123456') as archive:
            archive.writeall(ds[0], 'month.xlsx', )


def get_accounting_queque(data_set, context):
    context.queque_accounting = {
    "date": str(data_set[0][1]['datum']),
    "queue":[
        {
            "month": data_set[1][1]['datum'].strftime("%B"),
            "amount": abs(data_set[1][1]['i']),
            "left": None,
            "total": abs(data_set[1][1]['i']) * 2,
            "current": False,
            "next": False
        },
        {
            "month": data_set[0][1]['datum'].strftime("%B"),
            "amount": abs(data_set[0][1]['i']),
            "left": data_set[1][1]['g'],
            "total": abs(data_set[1][1]['i']) * 2 + abs(data_set[0][1]['i']),
            "current": False,
            "next": False
        },
        {
            "month": date.today().strftime("%B"),
            "amount": abs(data_set[0][1]['j']),
            "left": data_set[0][1]['g'],
            "total": None,
            "current": False,
            "next": False
        }
      ]
    }

def get_accounting_accountdata(data_set, context):
    context.account_data = []
    for ds in data_set:
        month = ds[1]['datum']-timedelta(days=31)

        context.account_data.insert(0, {
            "month": month.strftime("%B"),
            "updated": str(ds[1]['datum']),
            "account": 34504148,
            "changes_total": ds[1]['d'] + ds[1]['e'] + ds[1]['m'] + ds[1]['n'],
            "changes_detail": [
                {
                    "name": "Reconciliation change",
                    "value": ds[1]['d']
                },
                {
                    "name": "SPD Correction",
                    "value": ds[1]['e']
                },
                {
                    "name": "Broken technique",
                    "value": ds[1]['m']
                },
                {
                    "name": "Coach session",
                    "value": ds[1]['n']
                },

            ]
        })


def create_account_xlsx(context):
    wb_to = load_workbook('generator/account.xlsx')
    ws_to = wb_to.active

    data = context.account_data
    for i in range(len(data)):
        coef = 6*i
        ws_to[f'A{2 + coef}'] = data[i]['month']
        ws_to[f'E{2 + coef}'] = date.fromisoformat(data[i]['updated'])
        ws_to[f'C{2 + coef}'] = data[i]['changes_total']
        ws_to[f'C{3 + coef}'] = data[i]['changes_detail'][0]['value']
        ws_to[f'C{4 + coef}'] = data[i]['changes_detail'][1]['value']
        ws_to[f'C{5 + coef}'] = data[i]['changes_detail'][2]['value']
        ws_to[f'C{6 + coef}'] = data[i]['changes_detail'][3]['value']

    wb_to.save('base/data_set/account_data_template.xlsx')

def create_queue_xlsx(context):
    wb_to = load_workbook('generator/queue.xlsx')
    ws_to = wb_to.active

    data = context.queque_accounting['queue']
    for i in range(len(data)):
        ws_to[f'A{2 + i}'] = data[i]['month'] if data[i]['month'] != None else ''
        ws_to[f'B{2 + i}'] = data[i]['amount'] if data[i]['month'] != None else ''
        ws_to[f'C{2 + i}'] = data[i]['total'] if data[i]['month'] != None else ''
        ws_to[f'D{2 + i}'] = data[i]['left'] if data[i]['month'] != None else ''


    wb_to.save('base/data_set/account_queue_template.xlsx')

def make_accounting_precondition(hr_id, context):
    """get data_set"""
    data_set = random_fees_creator()
    """create .7z archives with fees"""
    create_7z(data_set, hr_id)
    """create accounting_queque template"""
    get_accounting_queque(data_set, context)
    """create account_data template"""
    get_accounting_accountdata(data_set, context)
    """create xlsx file account_data"""
    create_account_xlsx(context)
    """create xlsx file account queque"""
    create_queue_xlsx(context)























